from flask import Flask, render_template, request, send_from_directory, redirect, url_for
from openpyxl import load_workbook
from fpdf import FPDF
import os, re, xlrd

uploads_folder = os.path.join('src', 'uploads')
if not os.path.exists(uploads_folder):
 os.makedirs(uploads_folder)

def xlsx_to_txt(input0, output1):
 file_extension = os.path.splitext(input0)[1].lower()
 if file_extension == '.xlsx':
  workbook = load_workbook(input0, data_only=True)
  with open(output1, 'w', encoding='utf-8') as txt_file:
   for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(values_only=True):
     row_filtered = [str(cell) for cell in row if cell is not None]
     row_str = '\t'.join(row_filtered)
     txt_file.write(row_str + '\n')
 elif file_extension == '.xls':
  workbook = xlrd.open_workbook(input0)
  with open(output1, 'w', encoding='utf-8') as txt_file:
   for sheet_name in workbook.sheet_names():
    sheet = workbook.sheet_by_name(sheet_name)
    for row_idx in range(sheet.nrows):
     row = sheet.row_values(row_idx)
     row_filtered = [str(cell) for cell in row if cell is not None]
     row_str = '\t'.join(row_filtered)
     txt_file.write(row_str + '\n')
 else:
  raise ValueError("Formato de archivo no admitido. Use .xlsx o .xls.")

f0 = os.path.join('src', 'found')
db0 = os.path.join('src', 'database')
folder_files = ['html', 'pdf', 'txt']

for i in folder_files:
 create_folder=os.path.join(f0, i)
 if not os.path.exists(create_folder):
  os.makedirs(create_folder)

if not os.path.exists(db0):
  os.makedirs(db0)

folder_found=os.path.join('src', 'found', 'txt')
html=os.path.join('src', 'found', 'html')
save_files=[]

def one_complements(ci):
 complements1_html = f'''<!DOCTYPE html>
<html lang="es"><head>
<link rel="stylesheet" href="../../../static/css/styles.css">
<link rel="stylesheet" href="../../../static/css/font.css">
<link rel="shortcut icon" href="../../../static/icon/find.ico" type="image/x-icon">
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Texer</title>
</head>
<body>
<div id="container">
<h3 id="dni">{ci}</h3>
<div class="results-of-id">
<script src="../../../static/js/script.js"></script>
'''
 return complements1_html

def txt_to_pdf(ci):
 input_txt = os.path.join('src', 'found', 'txt', f'{ci}.txt')
 output_pdf = os.path.join('src', 'found', 'pdf', f'{ci}.pdf')
 pdf = FPDF()
 pdf.add_page()
 pdf.set_font("Arial", size=12)
 with open(input_txt, "r") as file:
  for line in file:
   write_in_file=' '.join(line.strip().split())
   pdf.multi_cell(0, 10, txt=write_in_file)
   pdf.ln(4)
 pdf.output(output_pdf)
 return output_pdf

def list_found():
 save_files.clear()
 list0 = os.listdir(folder_found)
 for i in list0:
  dni = i[:-4]
  save_files.append(dni)

list_found()
def save_in_file(file, file2, ci):
 rt1=os.path.join(folder_found, f'{ci}.txt')
 rt2=os.path.join(html, f'{ci}.html')
 if os.path.exists(rt1):
  os.remove(rt1)
 if os.path.exists(rt2):
  os.remove(rt2)
 for d in file:
  with open(rt1, 'a') as f:
   f.write(d)
   f.close()
 with open(rt2, 'a', encoding="UTF-8") as f:
  f.write(one_complements(ci))
  for d in file2:
   f.write(f'{d}')
  f.write('</div></body></html>')

found=[] ; found2=[]

def search(CI):
 found.clear() ; found2.clear()
 bd = os.path.join("src", "database")
 files = os.listdir(bd) ; number=1
 for file in files:
  ruta=os.path.join(bd, file)
  with open(ruta, "r", encoding="UTF-8") as f:
   ver=f.readlines()
   for result in ver:
    buscar=re.findall(CI, result)
    for d in buscar:
     if CI == d:
      write0=f"<div class='div-results'><label class='number'>[{number}]</label> {result.replace('\t', ' ')}.<br></div>"
      found.append(write0) ; found2.append(f'[{number}] {result.replace("\t", " ")}')
      number+=1;
 return found

app = Flask(__name__)
@app.route('/', methods=['GET', 'POST'])
def index():
 if request.method == 'POST':
  ci = request.form.get('ci')
  found0 = search(ci)
  save_in_file(found2, found, ci)
  list_found()
  return render_template('index.html', ci=found0, cedula=ci, sfile=save_files)
 return render_template('index.html', sfile=save_files, cedula=None)

@app.route('/generate_pdf/<ci>')
def generate_pdf(ci):
 pdf_path = txt_to_pdf(ci)
 return f"PDF generado: {pdf_path}"

@app.route('/found/html/<filename>')
def serve_html(filename):
 html_folder = os.path.join(app.root_path, 'src', 'found', 'html')
 return send_from_directory(html_folder, filename)

@app.route('/upload', methods=['POST'])
def upload_files():
 if 'files[]' not in request.files:
  return "No se seleccionaron archivos", 400
 files = request.files.getlist('files[]')
 for file in files:
  if file.filename == '':
   continue
  file_path = os.path.join('src', 'uploads', file.filename)
  file.save(file_path)
  output_txt = os.path.join('src', 'database', os.path.splitext(file.filename)[0] + '.txt')
  try:
   xlsx_to_txt(file_path, output_txt)
  except Exception as e:
   return f"Error al convertir {file.filename}: {str(e)}", 500
 return redirect(url_for('index'))

if __name__ == '__main__':
 app.run(debug=True)
